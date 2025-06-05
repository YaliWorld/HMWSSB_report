from flask import Flask, render_template, request, send_file, flash, redirect, url_for
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
import tempfile
from werkzeug.utils import secure_filename
import zipfile
from io import BytesIO

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Change this in production
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Allowed file extensions
ALLOWED_EXTENSIONS = {'csv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_and_format(file_path, output_file):
    """
    Reads a CSV file, verifies required columns, splits the data into
    GovtCAN Yes/No, writes them to an Excel file with two sheets,
    and then applies header styles and auto-adjusts column widths.
    """
    # Load dataset
    df = pd.read_csv(file_path)

    # Check required columns
    required_cols = ['ISGOVTCAN', 'DIVNCODE', 'LASTDEMAND']
    if not all(col in df.columns for col in required_cols):
        raise ValueError(f"Required columns {required_cols} do not exist in the dataset.")

    # Split into GovtCAN Yes and No, sorted by LASTDEMAND descending
    df_yes = df[df['ISGOVTCAN'] == 'Yes'].sort_values(by='LASTDEMAND', ascending=False)
    df_no = df[df['ISGOVTCAN'] == 'No'].sort_values(by='LASTDEMAND', ascending=False)

    # Write to Excel using xlsxwriter
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        df_yes.to_excel(writer, sheet_name='GovtCAN_Yes', index=False)
        df_no.to_excel(writer, sheet_name='GovtCAN_No', index=False)
    # ExcelWriter is now closed and file handles released

    # Open the file with openpyxl for styling
    wb = load_workbook(output_file)

    # Define style parameters
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    # Apply styles for each sheet
    for sheet_name in ['GovtCAN_Yes', 'GovtCAN_No']:
        ws = wb[sheet_name]
        for col_idx, cell in enumerate(ws[1], start=1):
            col_letter = get_column_letter(col_idx)
            cell.font = header_font
            cell.alignment = header_align

            # Highlight Column A and Column P headers if they exist
            if col_letter in ['A', 'P']:
                cell.fill = header_fill

            # Auto-adjust column width based on the maximum length in each column
            max_length = max(len(str(ws.cell(row=row, column=col_idx).value or "")) for row in range(1, ws.max_row + 1))
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_file)
    wb.close()  # Explicitly close the workbook

def process_combined_files(file_a_path, file_b_path, output_file):
    """
    Process both Category A and B files and create a combined Excel file
    """
    # Load both CSV files
    df_a = pd.read_csv(file_a_path)
    df_b = pd.read_csv(file_b_path)

    # Add the 'HCC Type' column to differentiate categories
    df_a["HCC Type"] = "A"
    df_b["HCC Type"] = "B"

    # Concatenate the dataframes
    combined_df = pd.concat([df_a, df_b], ignore_index=True)

    # Save as Excel file
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        combined_df.to_excel(writer, sheet_name='Combined', index=False)
    # ExcelWriter is now closed and file handles released

    # Format the combined Excel file using openpyxl
    wb_combined = load_workbook(output_file)
    ws_combined = wb_combined["Combined"]

    # Apply the same style parameters for the header row
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, cell in enumerate(ws_combined[1], start=1):
        col_letter = get_column_letter(col_idx)
        cell.font = header_font
        cell.alignment = header_align

        # Apply fill to Column A and Column P if present
        if col_letter in ['A', 'P']:
            cell.fill = header_fill

        # Auto-adjust column widths
        max_length = max(len(str(ws_combined.cell(row=row, column=col_idx).value or "")) for row in range(1, ws_combined.max_row + 1))
        ws_combined.column_dimensions[col_letter].width = max_length + 2

    wb_combined.save(output_file)
    wb_combined.close()  # Explicitly close the workbook

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/process_single', methods=['POST'])
def process_single():
    if 'file' not in request.files:
        flash('No file selected')
        return redirect(url_for('index'))
    
    file = request.files['file']
    category = request.form.get('category', 'A')
    
    if file.filename == '':
        flash('No file selected')
        return redirect(url_for('index'))
    
    if file and allowed_file(file.filename):
        temp_input_path = None
        temp_output_path = None
        
        try:
            # Create temporary input file
            temp_input = tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.csv')
            temp_input_path = temp_input.name
            file.save(temp_input_path)
            temp_input.close()  # Close the file handle
            
            # Create temporary output file
            temp_output = tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.xlsx')
            temp_output_path = temp_output.name
            temp_output.close()  # Close the file handle
            
            # Process the file
            process_and_format(temp_input_path, temp_output_path)
            
            # Read the output file into memory
            with open(temp_output_path, 'rb') as f:
                file_data = f.read()
            
            # Clean up temporary files
            try:
                if temp_input_path and os.path.exists(temp_input_path):
                    os.unlink(temp_input_path)
                if temp_output_path and os.path.exists(temp_output_path):
                    os.unlink(temp_output_path)
            except:
                pass  # Ignore cleanup errors
            
            # Send the file data from memory
            return send_file(
                BytesIO(file_data),
                as_attachment=True,
                download_name=f'Unbilled_CAT_{category}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        except Exception as e:
            # Clean up temporary files in case of error
            try:
                if temp_input_path and os.path.exists(temp_input_path):
                    os.unlink(temp_input_path)
                if temp_output_path and os.path.exists(temp_output_path):
                    os.unlink(temp_output_path)
            except:
                pass
            
            flash(f'Error processing file: {str(e)}')
            return redirect(url_for('index'))
    
    else:
        flash('Invalid file type. Please upload a CSV file.')
        return redirect(url_for('index'))

@app.route('/process_combined', methods=['POST'])
def process_combined():
    if 'file_a' not in request.files or 'file_b' not in request.files:
        flash('Both Category A and B files are required')
        return redirect(url_for('index'))
    
    file_a = request.files['file_a']
    file_b = request.files['file_b']
    
    if file_a.filename == '' or file_b.filename == '':
        flash('Both files must be selected')
        return redirect(url_for('index'))
    
    if (file_a and allowed_file(file_a.filename) and 
        file_b and allowed_file(file_b.filename)):
        
        temp_files = []
        
        try:
            # Create temporary input files
            temp_a = tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.csv')
            temp_a_path = temp_a.name
            temp_files.append(temp_a_path)
            file_a.save(temp_a_path)
            temp_a.close()
            
            temp_b = tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.csv')
            temp_b_path = temp_b.name
            temp_files.append(temp_b_path)
            file_b.save(temp_b_path)
            temp_b.close()
            
            # Create temporary output files
            temp_out_a = tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.xlsx')
            temp_out_a_path = temp_out_a.name
            temp_files.append(temp_out_a_path)
            temp_out_a.close()
            
            temp_out_b = tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.xlsx')
            temp_out_b_path = temp_out_b.name
            temp_files.append(temp_out_b_path)
            temp_out_b.close()
            
            temp_combined = tempfile.NamedTemporaryFile(mode='wb', delete=False, suffix='.xlsx')
            temp_combined_path = temp_combined.name
            temp_files.append(temp_combined_path)
            temp_combined.close()
            
            # Process the files
            process_and_format(temp_a_path, temp_out_a_path)
            process_and_format(temp_b_path, temp_out_b_path)
            process_combined_files(temp_a_path, temp_b_path, temp_combined_path)
            
            # Create a ZIP file with all outputs
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                zip_file.write(temp_out_a_path, 'Unbilled_CAT_A.xlsx')
                zip_file.write(temp_out_b_path, 'Unbilled_CAT_B.xlsx')
                zip_file.write(temp_combined_path, 'Unbilled_CAT_A_and_B_Combined.xlsx')
            
            zip_buffer.seek(0)
            
            # Clean up temporary files
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.unlink(temp_file)
                except:
                    pass
            
            return send_file(
                zip_buffer,
                as_attachment=True,
                download_name='processed_files.zip',
                mimetype='application/zip'
            )
        
        except Exception as e:
            # Clean up temporary files in case of error
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.unlink(temp_file)
                except:
                    pass
            
            flash(f'Error processing files: {str(e)}')
            return redirect(url_for('index'))
    
    else:
        flash('Invalid file type. Please upload CSV files only.')
        return redirect(url_for('index'))

